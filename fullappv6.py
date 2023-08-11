import cv2
import pytesseract as tess
from datetime import datetime
import numpy as np
from PIL import Image
import pandas as pd
import torch
import openpyxl
import time
import matplotlib.pyplot as plt


def load_model(model_path):
    # Load the pre-trained YOLOv5 model
    model = torch.hub.load('yolov5', 'custom', path=model_path, source='local')
    return model


def load_class_labels(labels_path):
    # Load the class labels for YOLOv5
    class_labels = []
    with open(labels_path, 'r') as f:
        class_labels = [line.strip() for line in f.readlines()]
    return class_labels


def preprocess_image(image):
    # Resize the image while maintaining aspect ratio
    width = 640
    height = int(image.shape[0] * width / image.shape[1])
    resized_image = cv2.resize(image, (width, height))

    # Convert image to RGB
    # image_rgb = cv2.cvtColor(resized_image, cv2.COLOR_BGR2RGB)
    image_grey = cv2.cvtColor(resized_image, cv2.COLOR_BGR2GRAY)
    # image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    # Convert image to PIL format
    image_pil = Image.fromarray(image_grey)
    return image_pil, resized_image


def perform_object_detection(model, image):
    # Perform object detection
    results = model(image)
    return results, np.array(image)



def extract_number_plates(results, image, confidence_threshold):
    # Process the detected number plates
    bounding_boxes = results.xyxy[0][:, :4].tolist()
    confidences = results.xyxy[0][:, 4].tolist()

    number_plates = []
    plate_bounding_boxes = []  # Store the bounding box coordinates

    image_width = image.shape[1]  # Width of the image
    image_height = image.shape[0]  # Height of the image

    for i in range(len(bounding_boxes)):
        confidence = confidences[i]

        if confidence > confidence_threshold:
            x1, y1, x2, y2 = map(int, bounding_boxes[i])
            x, y, width, height = x1 + 5, y1 - 5, x2 - x1 - 10, y2 - y1 + 10

            # Perform bounds checking
            if x < 0 or y < 0 or width <= 0 or height <= 0:
                print("Invalid region coordinates")
                continue

            # Convert the PIL Image to a NumPy array
            image_np = np.array(image)


            # Crop the number plate region from the image
            number_plate_region = image_np[y:y+height, x:x+width]

            # Convert the NumPy array to a PIL Image
            number_plate_image = Image.fromarray(number_plate_region)

            number_plates.append(number_plate_image)

            # Store the bounding box coordinates
            plate_bounding_boxes.append((x, y, width, height))



    return number_plates, plate_bounding_boxes


# def perform_ocr(image):
#
#     # Apply OCR on the image
#     number_plate_text = tess.image_to_string(image)
#     return number_plate_text

def perform_ocr(image):
    try:
        # Convert the PIL Image to a NumPy array
        image_np = np.array(image)

        # Apply OCR on the image
        number_plate_text = tess.image_to_string(image_np)
        print(f'text: {number_plate_text}')
        return number_plate_text
    except ValueError:
        print("Invalid image dimensions or data")
        return ""


def save_data_to_excel(data, file_path):
    # try:
    #     # Load the existing workbook
    #     workbook = openpyxl.load_workbook(file_path)
    #     # Select the first sheet
    #     sheet = workbook.active
    #     # Get the last row index in the sheet
    #     last_row = sheet.max_row
    #
    #     # Append the new data to the sheet
    #     for row in data:
    #         sheet.append(row)
    #
    #     # Save the workbook
    #     workbook.save(file_path)
    # except FileNotFoundError:
    #     # If the file doesn't exist, create a new workbook
    #     workbook = openpyxl.Workbook()
    #     # Select the first sheet
    #     sheet = workbook.active
    #
    #     # Write the data to the sheet
    #     number_plate_data = pd.DataFrame(data, columns=['Number Plate', 'Timestamp'])
    #     for row in data:
    #         sheet.append(row)
    #
    #     # Save the workbook
    #     workbook.save(file_path)
    try:
    # Load the existing workbook
        workbook = openpyxl.load_workbook(file_path)
        # Select the first sheet
        sheet = workbook.active
        # Get the last row index in the sheet
        last_row = sheet.max_row

        # Append the new data to the sheet
        for row in data:
            sheet.append(row)

        # Save the workbook
        workbook.save(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new DataFrame
        number_plate_data = pd.DataFrame(data, columns=['Number Plate', 'Timestamp'])
        #
        # # Save the data to the Excel file
        number_plate_data.to_excel(file_path, index=False)




def display(image, bounding_boxes):
    image_copy = image.copy()
    print(f'{bounding_boxes}')
    for (x, y, width, height) in bounding_boxes:
        x1, y1, x2, y2 = x, y, x + width, y + height
        cv2.rectangle(image_copy, (x1, y1), (x2, y2), (0, 255, 0), 2)
    # for (x1, y1, x2, y2) in bounding_boxes:
    #     cv2.rectangle(image_copy, (x1, y1), (x2, y2), (0, 255, 0), 2)

    cv2.imshow('Image', image_copy)
    cv2.waitKey(1)  #delay
    cv2.destroyAllWindows()


def process_images(model, class_labels, image_path, excel_file_path, confidence_threshold):
    image = cv2.imread(image_path)
    image_pil, imgo = preprocess_image(image)
    results, image = perform_object_detection(model, image_pil)
    number_plates, bounding_boxes = extract_number_plates(results, image, confidence_threshold)
    data = []
    number_plate_text = ''
    for number_plate in number_plates:
        number_plate_text = perform_ocr(number_plate)
        timestamp = datetime.now()
        data.append([number_plate_text, timestamp])
    save_data_to_excel(data, excel_file_path)
    display(image, bounding_boxes)

    print(bounding_boxes)
    print(number_plate_text)


# def process_video(model, class_labels, video_path, excel_file_path):
#     data = []
#     bounding_boxes = []
#
#     cap = cv2.VideoCapture(video_path)
#
#     while cap.isOpened():
#         ret, frame = cap.read()
#
#         if not ret:
#             break
#
#         # Preprocess the frame
#         frame_pil = preprocess_image(frame)
#
#         # Perform object detection
#         results, frame = perform_object_detection(model, frame_pil)
#
#         # Extract the number plates and bounding box coordinates
#         number_plates, boxes = extract_number_plates(results, frame)
#
#         for number_plate in number_plates:
#             number_plate_text = perform_ocr(number_plate)
#             timestamp = datetime.now()
#             data.append([number_plate_text, timestamp])
#             bounding_boxes.append(number_plate)  # Add the bounding box coordinates
#
#     save_data_to_excel(data, excel_file_path)
#     display(frame, bounding_boxes)
#     cap.release()

# def process_video(model, class_labels, video_path, confidence_threshold, excel_file_path):
#     data = []
#     bounding_boxes = []
#
#     cap = cv2.VideoCapture(video_path)
#
#     while cap.isOpened():
#         ret, frame = cap.read()
#
#         if not ret:
#             break
#
#         # Preprocess the frame
#         frame_pil = preprocess_image(frame)
#
#         # Perform object detection
#         results, frame = perform_object_detection(model, frame_pil)
#
#         # Extract the number plates and bounding box coordinates
#         number_plates, boxes = extract_number_plates(results, frame, confidence_threshold)
#
#         for number_plate in number_plates:
#             number_plate_text = perform_ocr(number_plate)
#             timestamp = datetime.now()
#             data.append([number_plate_text, timestamp])
#             bounding_boxes.append(number_plate)  # Add the bounding box coordinates
#
#     save_data_to_excel(data, excel_file_path)
#     display(frame, bounding_boxes)
#     cap.release()

def process_video(model, class_labels, video_path, confidence_threshold, excel_file_path):
    data = []
    bounding_boxes = []

    cap = cv2.VideoCapture(video_path)

    # Read the first frame
    ret, frame = cap.read()

    while ret:
        # Preprocess the frame
        frame_pil, imgo = preprocess_image(frame)

        # Perform object detection
        results, frame = perform_object_detection(model, frame_pil)

        # Extract the number plates and bounding box coordinates
        number_plates, boxes = extract_number_plates(results, frame, confidence_threshold)

        for number_plate in number_plates:
            number_plate_text = perform_ocr(number_plate)
            timestamp = datetime.now()
            data.append([number_plate_text, timestamp])
            bounding_boxes.append(number_plate)  # Add the bounding box coordinates

        # Display the frame
        display(frame, bounding_boxes)

        # Read the next frame
        ret, frame = cap.read()

    save_data_to_excel(data, excel_file_path)
    cap.release()



def process_video_with_display(model, class_labels, video_path, confidence_threshold, excel_file_path):
    data = []
    bounding_boxes = []

    cap = cv2.VideoCapture(video_path)
    fps_start_time = time.time()
    fps_counter = 0

    while cap.isOpened():
        ret, frame = cap.read()

        if not ret:
            break

        # Preprocess the frame
        frame_pil, img = preprocess_image(frame)

        # Perform object detection
        results, frame = perform_object_detection(model, frame_pil)

        # Extract the number plates and bounding box coordinates
        number_plates, boxes = extract_number_plates(results, frame, confidence_threshold)
        if number_plates:
            for number_plate in number_plates:
                number_plate_text = perform_ocr(number_plate)
                timestamp = datetime.now()
                data.append([number_plate_text, timestamp])
                bounding_boxes.append(number_plate)  # Add the bounding box coordinates

        # Display the frame with bounding boxe s
        display(img, boxes)
        # display(o_frame, boxes)

        # Calculate frame rate
        fps_counter += 1
        if (time.time() - fps_start_time) > 1:
            fps = fps_counter / (time.time() - fps_start_time)
            print("Frame Rate: {:.2f} fps".format(fps))
            fps_start_time = time.time()
            fps_counter = 0

    save_data_to_excel(data, excel_file_path)
    cap.release()






def process_live_video(model, confidence_threshold, excel_file_path):
    data = []
    bounding_boxes = []

    cap = cv2.VideoCapture(0)

    while cap.isOpened():
        ret, frame = cap.read()

        if not ret:
            break

        # Preprocess the frame
        frame_pil, imgo = preprocess_image(frame)

        # Perform object detection
        results, frame = perform_object_detection(model, frame_pil)

        # Extract the number plates and bounding box coordinates
        number_plates, boxes = extract_number_plates(results, frame, confidence_threshold)

        for number_plate in number_plates:
            number_plate_text = perform_ocr(number_plate)
            timestamp = datetime.now()
            data.append([number_plate_text, timestamp])
            bounding_boxes.append(number_plate)  # Add the bounding box coordinates

        cv2.imshow('Live Video', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    save_data_to_excel(data, excel_file_path)
    cap.release()
    cv2.destroyAllWindows()


if __name__ == '__main__':
    # Paths and configurations
    model_path = 'ANPRwv5.pt'
    labels_path = 'labels.txt'
    excel_file_path = 'number_plate_data.xlsx'
    confidence_threshold = 0.2

    tess.pytesseract.tesseract_cmd = r'D:\download\PyTesseract\tesseract.exe'

    # Load the pre-trained YOLOv5 model
    model = load_model(model_path)

    # Load the class labels for YOLOv5
    class_labels = load_class_labels(labels_path)

    # choose the environment to run
    running = 'img'  #img, vid, live


    if running == 'img':
        # Use case 1: Process a single image
        # image_path = 'resx/pics/ef_00 (1).jpg'
        image_path = 'resx/pics/2.jpg'
        process_images(model, class_labels, image_path, excel_file_path, confidence_threshold)

    elif running == 'vid':
        # Use case 2: Process a video file
        video_path = 'resx/pics/VID_003.mp4'
        # process_video(model, class_labels, video_path, excel_file_path)
        # process_video(model, class_labels, video_path, confidence_threshold, excel_file_path)
        process_video_with_display(model, class_labels, video_path, confidence_threshold, excel_file_path)


    elif running == 'live':
        # Use case 3: Live video processing
        process_live_video(model, confidence_threshold, excel_file_path)

    else:
        print('Not Valid')








